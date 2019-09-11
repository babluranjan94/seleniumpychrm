import openpyxl


testcase_list=[]
path = r"D:\xstudy\Test_Cases_Xtudy.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

for row in sheet.iter_rows(min_row=2, max_col=cols, max_row=rows):
    for cell in row:
        testcase_list.append(cell.value)
    #if testcase_list[2] == 'Y' and testcase_list[3]=='LOGIN':
    print (testcase_list)
    del testcase_list[:]



print (rows)
print (cols)



