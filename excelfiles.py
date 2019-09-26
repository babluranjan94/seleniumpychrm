import openpyxl
import Xtudylogin

propertydict = {} #dictonary of the property file that contains all the Xpath value
testcase_list = [] # list of test cases from test case sheet
test_data_comand = []# getting the command from the 1 row of test data sheet
test_data_xpath =[] # getting the xpath from the second row of test data sheet
test_data_value = []

with open(r'C:\Users\Bablu\PycharmProjects\seleniumpychrm\property.txt') as propertyfile:
    for line in propertyfile:
        line = line.strip()
        x = line.split('==')
        a = x[0]
        b = x[1]
        propertydict[a] = b

propertyfile.close()






test_case_path = r"C:\Users\Bablu\PycharmProjects\seleniumpychrm\Test_Cases_Xtudy.xlsx"
test_data_path=r"C:\Users\Bablu\PycharmProjects\seleniumpychrm\Test_Data.xlsx"

workbook = openpyxl.load_workbook(test_case_path)
test_data_workbook=openpyxl.load_workbook(test_data_path)

sheet = workbook["TEST_EXECUTION"]

rows = sheet.max_row
cols = sheet.max_column


for row in sheet.iter_rows(min_row=2, max_col=cols, max_row=rows):
    for cell in row:
        testcase_list.append(cell.value) # getting the test case from the test case sheet

    if testcase_list[2] == 'Y' and testcase_list[3] == 'LOGIN': # for the login
        print testcase_list
        print Xtudylogin.applogin(propertydict['APPURL'], propertydict['LOGINID'], propertydict['PASSWORD'],propertydict['XUSERINPUT'],
                                    propertydict['XUSERPASS'], propertydict['XLOGINBUTT'])

        del testcase_list[:]
    elif testcase_list[2] == 'Y' and testcase_list[3] == 'LOGOUT':
        print testcase_list
        Xtudylogin.logout()
        del testcase_list[:]
    elif testcase_list[2] == 'Y':
        '''the control is shifting from test case sheet to 
        test data sheet'''

        test_data_sheet = test_data_workbook[testcase_list[3]]
        for row in test_data_sheet.iter_rows(min_row=1, max_row=1):  # getting the command from the first row of test data sheet
            for cell in row:
                test_data_comand.append(cell.value)
            print test_data_comand
        for row in test_data_sheet.iter_rows(min_row=2, max_row=2): #getting the Xpath Keys for the propertydict
            for cell in row:
                test_data_xpath.append(cell.value)
            print test_data_xpath

            '''checking the data now'''
        for row in test_data_sheet.iter_rows(min_row=3):  # getting the value to enter or verify
            for cell in row:
                test_data_value.append(cell.value)
            print test_data_value
            print len(test_data_value)
            for i in range(1, len(test_data_value)):
                print "no. of iteration = ", i
                if test_data_comand[i] == "CLICK" and test_data_value[i] != "#SKIP#":
                    print test_data_comand[i]
                    print propertydict[test_data_xpath[i]]
                    Xtudylogin.CLICK(propertydict[test_data_xpath[i]])
                elif test_data_comand[i] == 'DROP_DOWN'and test_data_value[i] != "#SKIP#":
                    print test_data_xpath[i]
                    Xtudylogin.dropdown(propertydict[test_data_xpath[i]], test_data_value[i])
                elif test_data_comand[i] == 'SCROLL' and test_data_value[i] != "#SKIP#":
                    Xtudylogin.down()
                elif test_data_comand[i] == 'WAIT'and test_data_value[i] != "#SKIP#":
                    print test_data_value[i]
                    Xtudylogin.Wait_sec(1)
                elif test_data_comand[i] == "SETTEXT" and test_data_value[i] != "#SKIP#":
                    print test_data_value[i]
                    print propertydict[test_data_xpath[i]]
                    Xtudylogin.SETTEXT(propertydict[test_data_xpath[i]], test_data_value[i])
                elif test_data_comand[i] == "DATE" and test_data_value[i] != "#SKIP#":
                    print test_data_value[i]
                    print propertydict[test_data_xpath[i]]
                    Xtudylogin.SET_DATE(propertydict[test_data_xpath[i]], test_data_value[i])
                elif test_data_comand[i] == "GETTEXT" and test_data_value[i] != "#SKIP#":
                    if Xtudylogin.GET_TEXT(propertydict[test_data_xpath[i]])== test_data_value[i]:
                        print ("pass")
                    else:
                        print ("fail")

                else:
                    print"work in progress"
            del test_data_value[:]
        del testcase_list[:]

    else:
        print testcase_list
        print("work on it")
        del testcase_list[:]
