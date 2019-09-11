import  openpyxl

wb = openpyxl.load_workbook(r"D:\xstudy\Test_Cases_Xtudy.xlsx") #obtaining the whole test case sheet
wb2 = openpyxl.load_workbook(r"D:\xstudy\Test_Data.xlsx")

ws = wb["TEST_EXECUTION"] # opening the sheet

test_case_list = [] # creating the list to execute
test_data_comand = []# getting the command
test_data_xpath =[] # getting the xpath
test_data_value = []
max_row= ws.max_row
max_cols = ws.max_column

print(max_row)
print (max_cols)


for row in ws.iter_rows(min_row=2, max_col=max_cols, max_row=max_row):
    for cell in row:
        test_case_list.append(cell.value)
    if test_case_list[2] == 'Y' and test_case_list[3] == 'ADD_STUDENT':
        ws2=wb2["ADD_STUDENT"]
        for row in ws2.iter_rows( min_row=1,max_row=1):
            for cell in row:
                test_data_comand.append(cell.value)
            print test_data_comand
        for row in ws2.iter_rows( min_row=2, max_row=2):
            for cell in row:
                test_data_xpath.append(cell.value)
            print test_data_xpath
        for row in ws2.iter_rows(min_row=3):
            for cell in row:
                test_data_value.append(cell.value)
            print test_data_value

        del test_case_list[:]
    else:
        del test_case_list[:]






